import softest
import logging
import inspect
from openpyxl import Workbook, load_workbook
from datetime import datetime
import csv

# anything independent of the driver reference goes to the utilities
class Utils(softest.TestCase):
    def assertListItemText(self, list, value):
        for stop in list:
            # print(f"Filtered flight text: {stop.text}")
            # #print("The text is:"+stop.text)
            # #assert stop.text == value
            # assert stop.text.strip().lower() == value.strip().lower()
            # print("assert passed")

            actual_text = stop.text.strip().lower()
            expected_text = value.strip().lower()
            print(f"Filtered flight text: '{actual_text}' (Expected: '{expected_text}')")
            # Using soft assertion properly
            self.soft_assert(self.assertEqual, actual_text, expected_text,
                             f"Assertion failed! Expected '{expected_text}', but got '{actual_text}'")

            # Ensures all soft assertions are checked
        self.assert_all()

    def custum_logger(logLevel = logging.DEBUG):
        # Set class/method name from where it is called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # create console handler or file handler to set the log level
        fh = logging.FileHandler("automation.log", mode="w")
        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(filename)s - %(levelname)s : %(message)s ', datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        # row count
        row_ct = sh.max_row
        # column count
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):  # Skip header row
            row = []
            for j in range(1, col_ct + 1):
                cell_value = sh.cell(row=i, column=j).value

                # ✅ Convert datetime object to string format "28 March 2025"
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime("%d %B %Y")

                row.append(cell_value)

            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        # Create an empty list
        datalist= []

        #Open CSV file
        csvdata = open(filename,"r")

        # Create a csv reader
        reader = csv.reader(csvdata)

        # Skip the header
        next(reader)

        # add csv rows to empty list
        for rows in reader:
            datalist.append(rows)

        return datalist